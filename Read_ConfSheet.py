
import openpyxl as pxl
import csv
from datetime import datetime

def find_value_in_row(Config_sheet, start_row, label, max_column):
    """Find a value in a row after a label cell"""
    for col in range(1, max_column + 1):
        cell = Config_sheet.cell(row=start_row, column=col)
        if cell.value and label.lower() in str(cell.value).lower():
            # Found label, now get the value in the next non-empty cell
            value_cell = Config_sheet.cell(row=start_row, column=col+1)
            attempts = 0
            while not value_cell.value and attempts < 10:
                value_cell = Config_sheet.cell(row=start_row, column=value_cell.column+1)
                attempts += 1
            return value_cell.value if value_cell.value else None
    return None

def find_article_numbers(Config_sheet, title_row, max_column, search_range=20):
    """Find finished part and machining part article numbers near the title"""
    finished_article = None
    machining_article = None
    
    # Search in rows around the title (before and after)
    for row_offset in range(-5, search_range):
        current_row = title_row + row_offset
        if current_row < 1:
            continue
            
        for col in range(1, max_column + 1):
            cell = Config_sheet.cell(row=current_row, column=col)
            cell_value = str(cell.value).lower() if cell.value else ""
            
            # Look for article number patterns
            if "article" in cell_value or "art." in cell_value or "art:" in cell_value:
                # Get the value in the next cell
                value_cell = Config_sheet.cell(row=current_row, column=col+1)
                attempts = 0
                while not value_cell.value and attempts < 5:
                    value_cell = Config_sheet.cell(row=current_row, column=value_cell.column+1)
                    attempts += 1
                
                article_value = value_cell.value
                
                # Determine if it's finished or machining part based on context
                # Check a few cells before for context keywords
                context = ""
                for ctx_col in range(max(1, col-3), col):
                    ctx_cell = Config_sheet.cell(row=current_row, column=ctx_col)
                    if ctx_cell.value:
                        context += str(ctx_cell.value).lower() + " "
                
                if "machin" in context:
                    if not machining_article:
                        machining_article = article_value
                elif "finished" in context or "fertig" in context:
                    if not finished_article:
                        finished_article = article_value
                else:
                    # If no specific context, assign to finished first, then machining
                    if not finished_article:
                        finished_article = article_value
                    elif not machining_article:
                        machining_article = article_value
    
    return finished_article, machining_article

def Readsheet(): 
    ###read file path from user input, load sheet 
    #file_path = input("Enter the path to your Excel file: ")
    file_path = "C:/Users/cmerkle/OneDrive - IDEX Corporation/Desktop/Configsheet_Project/CS-BRAC055U007.xlsx"
    #file_path = file_path.replace("\\", "/")  # Replace backslashes with forward slashes

    Config_file = pxl.load_workbook(file_path, data_only=True)

    # Get the last sheet in the workbook 
    Config_sheet = Config_file.worksheets[-1]
    #print(Config_sheet.title)
    #print(f"Max rows: {Config_sheet.max_row}, Max columns: {Config_sheet.max_column}")
    max_column = Config_sheet.max_column

    #get Produkt titles and article numbers
    parts_data = []
    partcount = 1       

    for row in Config_sheet.iter_rows():
        #Find Title cell, then search for the next non-empty cell in the same row to get the title value 
      
        for cell in row:
            if cell.value and "Title" in str(cell.value):
                titlecell = Config_sheet.cell(row=cell.row, column=cell.column+1)

                while not titlecell.value: 
                    if (titlecell.column) > max_column:
                        print(f"No title found for part {partcount} at row {cell.row}.")
                        break
                    
                    titlecell = Config_sheet.cell(row=cell.row, column=titlecell.column+1)
                
                title_value = titlecell.value
                print(f"\nPart {partcount}: {title_value}")
                
                # Find article numbers for this part
                finished_article, machining_article = find_article_numbers(
                    Config_sheet, cell.row, max_column
                )
                
                print(f"  Finished Part Article: {finished_article if finished_article else 'Not found'}")
                print(f"  Machining Part Article: {machining_article if machining_article else 'Not found'}")
                
                parts_data.append({
                    'Part_Number': partcount,
                    'Title': title_value,
                    'Finished_Part_Article': finished_article,
                    'Machining_Part_Article': machining_article
                })
                
                partcount += 1
                
    if not parts_data:
        print("No cells with 'Title' found.")
        return
    
    # Write to CSV file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Parts_List_{timestamp}.csv"
    
    with open(output_filename, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['Part_Number', 'Title', 'Finished_Part_Article', 'Machining_Part_Article']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        writer.writerows(parts_data)
    
    print(f"\n✓ Data exported to {output_filename}")
    print(f"Total parts extracted: {len(parts_data)}")
    
    return parts_data     

if __name__ == "__main__": 
    Readsheet()